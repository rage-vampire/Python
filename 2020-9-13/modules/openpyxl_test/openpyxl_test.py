#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @File  : openpyxl_test.py
# @Author: Lizi
# @Date  : 2020/10/28

import openpyxl
import faker

# excel = openpyxl.Workbook()

excel = openpyxl.load_workbook('./user_info.xlsx')           # 获取一个已有的excel表格
# print(excel)
excel.create_sheet('user', 0)                                # 在第一个位置创建一个名为user的sheet单元表
active_sheet = excel.active                                  # 获取活跃的sheet单元表
print(active_sheet)
# print(excel.worksheets)                                      # 获取所有worksheet对象
# print(excel.read_only)                                       # 判断是否以read_only模式打开Excel文档
# print(excel.encoding)                                        # 获取文档的字符集编码
# print(excel.properties)                                      # 获取文档的元数据，如标题，创建者，创建日期等
print(excel.sheetnames)                                         # 获取表格中所有sheet单元表（列表）

sheet = excel['Sheet1']                                        # 获取名为user的worksheet对象
# active_sheet.append(['name', 'tel', 'address'])

print(sheet)
# print(sheet.title)         # 获取单元表的名字
# print(sheet.dimensions)    # 获取表格的区域（包含数据的部分）
# print(sheet.max_row)       # 获取表格最大行（包含数据的部分）
# print(sheet.min_row)       # 获取表格最小行（包含数据的部分）
# print(sheet.max_column)    # 获取表格最大列（包含数据的部分）
# print(sheet.min_column)    # 获取表格最小行（包含数据的部分）


# 读单个单元格
print(sheet['A1'].value)
print(sheet.cell(row=1, column=2).value)


# 读多个单元格
for i in sheet['A1':'C3']:
    for cell in i:
        print(cell.value)


# 按列获取单元格(Cell对象) - 生成器
# for column in sheet.columns:
#     print(column)
#     for cell in column:
#         print(cell.value)

# for cell in list(sheet.columns)[0]:
#     print(cell.value)
#     print(cell.coordinate)


# sheet.values 按行获取表格的内容-生成器
# for i in sheet.values:
#     print(i)

# for row in sheet.iter_rows('A1:C2'):
#     for cell in row:
#         print(cell.value)