#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @File  : faker_test.py
# @Author: Lizi
# @Date  : 2020/10/27

import os
from faker import Faker
import openpyxl


def create_excel():
    faker = Faker(locale='zh_CN')
    excel = openpyxl.Workbook()                    # 创建一个新的excel表格
    # excel = openpyxl.load_workbook('./user_info.xlsx')    # 获取一个已有的excel表格
    all_sheet = excel.sheetnames                     # 获取所有sheet，以列表的方式返回

    '''判断user_sheet表格是否已经存在，如果已经存在则直接获取user_sheet表格'''
    if 'user_sheet' not in all_sheet:
        excel.create_sheet('user_sheet', 0)           # 创建一个新的sheet表格,0代表第一个表格
        user_sheet = excel['user_sheet']              # 获取user_sheet表格
        # print(type(user_sheet))
    else:
        user_sheet = excel['user_sheet']

    user_sheet.append(['name', 'tel', 'address'])    # 在user_sheet表格中添加表格title,user_sheet是一个列表类型
    for i in range(100):
        user_sheet.append([faker.name(), faker.phone_number(), faker.address()])   # 通过faker库获取用户信息，添加到user_sheet表格中
    excel.save('./user_info.xlsx')                        # 保存excel文件
    # print(user_sheet)


def find_info():
    excel = openpyxl.load_workbook('user_info.xlsx')         # 获取一个已有的excel表格
    user_sheet = excel['user_sheet']                         # 获取名为user_sheet的sheet单元表
    name = input('请输入姓名：')
    # sex = input('请输入性别：')
# ----------------------------------------------------------------------------------------------------------------------
    # for cell in list(user_sheet.columns)[0]:                 # user_sheet.columns是一个生成器类型，不能使用索引，将user_sheet.columns转换成列表类型，可通过索引方式获取指定列的数据
    #     user_info = []
    #     for info in list(user_sheet.rows)[cell.row - 1]:     # cell.row 获取当前单元格的行
    #         user_info.append(info.value)
    #     if name in user_info:
    #         print('用户信息：', user_info)
    #         break
    # else:
    #     print('没有找到{}信息'.format(name))
# ----------------------------------------------------------------------------------------------------------------------
    for info in user_sheet.values:            # user_sheet.values安行获取表格的内容（生成器），需使用迭代获取数据
        if name in info:
            print(info)
            break
    else:
        print('没有找到{}信息'.format(name))


if __name__ == '__main__':
    create_excel()
    find_info()

# for column in user_sheet.columns:              # user_sheet.columns是一个生成器类型，不能使用索引
#     for cell in column:
#         print('所有单元格的数据：', cell.value)